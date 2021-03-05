import pandas as pd
from openpyxl import styles
from openpyxl.utils import get_column_letter


def dicts_to_excel(filename, data: list[dict], sheet_name='Sheet1', show_extra_cols=False) -> None:
    if str(filename)[-5:] != '.xlsx':
        filepath = str(filename) + '.xlsx'
    else:
        filepath = str(filename)
    writer = pd.ExcelWriter(filepath, engine='openpyxl')
    df = pd.DataFrame(data)
    df.to_excel(writer, sheet_name=sheet_name)

    worksheet = writer.book.active

    # set col width to match text, and apply text wrap to those who need it
    for i, col_width in enumerate(__get_col_widths(df)):
        if worksheet[f'{get_column_letter(i + 1)}1'].value in ['Line Appearances', 'Busy Lamp Fields', 'Call Pickup Group']:
            worksheet.column_dimensions[get_column_letter(i + 1)].width = len(worksheet[f'{get_column_letter(i + 1)}1'].value) + 1
            for row in range(2, len(data) + 1, 1):
                worksheet.cell(row=row,column=i+1).alignment = styles.Alignment(wrap_text=True)
        else:
            worksheet.column_dimensions[get_column_letter(i+1)].width = col_width + 1

    # hide unneeded cols
    if not show_extra_cols:
        cols_to_hide = 7
        total_cols = len(data[0].keys()) + 1
        for n in range((total_cols - cols_to_hide) + 1, total_cols + 1):
            worksheet.column_dimensions[get_column_letter(n)].hidden = True

    writer.save()


def __get_col_widths(dataframe):
    # First we find the maximum length of the index column
    idx_max = max([len(str(s)) for s in dataframe.index.values] + [len(str(dataframe.index.name))])
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
    return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]
